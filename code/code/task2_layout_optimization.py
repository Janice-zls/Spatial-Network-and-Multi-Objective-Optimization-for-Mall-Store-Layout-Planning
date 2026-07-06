

import csv
import math
import random
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import matplotlib.pyplot as plt

from plot_style import (
    BLUE_DARK,
    BLUE_LIGHT,
    BLUE_MID,
    COLOR_WHITE,
    GREEN_DARK,
    GREEN_LIGHT,
    GREEN_MID,
    PURPLE_DARK,
    PURPLE_LIGHT,
    PURPLE_MID,
    TEXT_DARK,
    cmap_from,
    place_legend,
    save_figure,
    setup_matplotlib_style,
    short_label,
)

ROOT = Path("SJMMA/Problem_B")
PROJECT_DIR = Path("SJMMA/2026_SJMMA")
FIGURE_DIR = PROJECT_DIR / "figures"
WORKBOOK_PATH = ROOT / "guangzhou_taikoo_hui_store_directory_integrated.xlsx"
NODE_FLOW_CSV = ROOT / "task1_node_flow_by_hour.csv"

COMPARISON_CSV = ROOT / "task2_layout_comparison.csv"
TRACE_CSV = ROOT / "task2_optimization_trace.csv"
SUMMARY_CSV = ROOT / "task2_summary.csv"

FIG_CONVERGENCE = FIGURE_DIR / "task2_optimization_convergence.png"
FIG_REVENUE_COMPARE = FIGURE_DIR / "task2_revenue_before_after.png"
FIG_GAIN_TOP10 = FIGURE_DIR / "task2_store_gain_top10.png"
FIG_MIGRATION = FIGURE_DIR / "task2_floor_migration_matrix.png"
FIG_CHANGE_TYPE = FIGURE_DIR / "task2_layout_change_by_type.png"
FIG_MULTI_SEED = FIGURE_DIR / "task2_multiseed_convergence.png"
FIG_GAIN_DISTRIBUTION = FIGURE_DIR / "task2_revenue_gain_distribution.png"
FIG_FLOOR_HEAT = FIGURE_DIR / "task2_before_after_floor_heatmap.png"
FIG_CATEGORY_BALANCE = FIGURE_DIR / "task2_category_balance.png"
FIG_CONSTRAINT = FIGURE_DIR / "task2_constraint_diagnostics.png"
FIG_OPT_DASHBOARD = FIGURE_DIR / "task2_optimization_dashboard.png"
FIG_LAYOUT_DASHBOARD = FIGURE_DIR / "task2_layout_diagnostics_dashboard.png"

HOURS = list(range(10, 22))
FLOORS = ["M", "MU", "L1", "L2", "L3"]
RANDOM_SEED = 20260501


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def classify_store_type(source_type: str, category_text: str, node_category: str, brand: str) -> str:
    text = f"{source_type} {category_text} {node_category} {brand}".lower()
    dining_words = ["美食", "餐饮", "风味", "烘焙", "咖啡", "甜品", "茶", "酒家", "炉端", "越南", "泰国"]
    anchor_words = ["主力店", "超市", "美食广场", "anchor", "olé", "ole", "方所"]
    if source_type == "美食" or any(word in text for word in dining_words):
        return "dining"
    if node_category == "Anchor_Retail" or any(word in text for word in anchor_words):
        return "anchor"
    return "retail"


def store_price(store_type: str, category_text: str) -> float:
    if store_type == "dining":
        if "美食广场" in category_text or "休闲" in category_text:
            return 55.0
        if "中式" in category_text or "亚洲" in category_text or "西式" in category_text:
            return 130.0
        return 75.0
    if store_type == "anchor":
        return 95.0
    if "珠宝" in category_text or "钟表" in category_text:
        return 1800.0
    if "女装" in category_text or "男装" in category_text:
        return 680.0
    if "鞋" in category_text or "包" in category_text:
        return 850.0
    if "美容" in category_text or "保健" in category_text:
        return 320.0
    return 420.0


def base_attraction(store_type: str, node_category: str, size_rating: int) -> float:
    if store_type == "anchor":
        base = 4.0
    elif store_type == "dining":
        base = 2.7
    else:
        base = 1.9
    if node_category == "Retail_Combined":
        base += 0.5
    if node_category == "Retail_Kiosk":
        base -= 0.2
    return max(0.3, base + 0.25 * max(1, size_rating))


def conversion_rate(store_type: str, hour: int) -> float:
    if store_type == "dining":
        base = 0.28
        if hour in (12, 13, 18, 19):
            base += 0.22
        return base
    if store_type == "anchor":
        return 0.32 if 11 <= hour <= 20 else 0.25
    base = 0.13
    if hour in (15, 16, 17, 20):
        base += 0.04
    return base


def time_multiplier(store_type: str, hour: int) -> float:
    if store_type == "dining":
        if hour in (12, 13):
            return 2.1
        if hour in (18, 19):
            return 2.4
        if hour in (20, 21):
            return 1.4
        return 1.0
    if store_type == "anchor":
        return 1.35 if 11 <= hour <= 20 else 1.1
    if hour in (15, 16, 17, 20):
        return 1.35
    return 1.0


def load_stores_and_positions() -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, str]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["整合主表"]
    headers = [cell.value for cell in ws[1]]

    stores: dict[str, dict[str, Any]] = {}
    positions: dict[str, dict[str, Any]] = {}
    real_layout: dict[str, str] = {}
    duplicate_count: Counter[str] = Counter()

    for excel_row, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, row_values))
        brand = clean_text(row.get("店铺/品牌"))
        node_id = clean_text(row.get("node_id"))
        if not brand or not node_id:
            continue

        floor = clean_text(row.get("楼层"))
        node_category = clean_text(row.get("节点类别"))
        category_text = clean_text(row.get("类别"))
        source_type = clean_text(row.get("来源类型"))
        size_rating = int(row.get("size_rating") or 1)
        store_type = classify_store_type(source_type, category_text, node_category, brand)

        duplicate_count[node_id] += 1
        position_id = f"{node_id}@{duplicate_count[node_id]}"
        store_id = f"S{excel_row}"

        positions[position_id] = {
            "position_id": position_id,
            "node_id": node_id,
            "floor": floor,
            "node_category": node_category,
            "size_rating": size_rating,
            "slot": clean_text(row.get("标准铺位号")),
        }
        stores[store_id] = {
            "store_id": store_id,
            "brand": brand,
            "original_position": position_id,
            "original_node": node_id,
            "original_floor": floor,
            "store_type": store_type,
            "category_text": category_text,
            "price": store_price(store_type, category_text),
            "need_size": size_rating,
            "attraction": base_attraction(store_type, node_category, size_rating),
        }
        real_layout[position_id] = store_id

    return stores, positions, real_layout


def load_node_flow(positions: dict[str, dict[str, Any]]) -> dict[int, dict[str, float]]:
    node_flow: dict[int, dict[str, float]] = {hour: defaultdict(float) for hour in HOURS}
    if NODE_FLOW_CSV.exists():
        with NODE_FLOW_CSV.open("r", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                hour = int(row["hour"])
                if hour in node_flow:
                    node_flow[hour][row["node_id"]] = float(row["flow"])
        return node_flow

    for hour in HOURS:
        for pos in positions.values():
            base = 10 + 4 * pos["size_rating"]
            if pos["node_category"] == "Anchor_Retail":
                base += 15
            node_flow[hour][pos["node_id"]] = base
    return node_flow


def is_assignment_feasible(store: dict[str, Any], position: dict[str, Any]) -> bool:
    return store["need_size"] <= position["size_rating"] + 2


def position_score(
    store: dict[str, Any],
    position: dict[str, Any],
    node_flow: dict[int, dict[str, float]],
) -> float:
    if not is_assignment_feasible(store, position):
        return -1_000_000.0

    value = 0.0
    flow_bonus = {"L1": 1.12, "M": 1.08, "MU": 1.03, "L2": 0.95, "L3": 0.88}.get(position["floor"], 1.0)
    if position["node_category"] == "Anchor_Retail" and store["store_type"] == "anchor":
        flow_bonus += 0.24
    if position["node_category"] == "Retail_Kiosk" and store["store_type"] == "retail":
        flow_bonus += 0.08
    if position["node_category"] == "Retail_Combined":
        flow_bonus += 0.10

    size_gap = abs(position["size_rating"] - store["need_size"])
    area_fit = max(0.72, 1.0 - 0.055 * size_gap)
    anchor_spillover = 1.10 if store["store_type"] in {"dining", "retail"} and position["size_rating"] >= 3 else 1.0
    category_bonus = 1.0
    if store["store_type"] == "dining" and position["floor"] in {"M", "MU", "L3"}:
        category_bonus += 0.08
    if store["store_type"] == "retail" and position["floor"] in {"L1", "L2"}:
        category_bonus += 0.06

    for hour in HOURS:
        flow = node_flow[hour].get(position["node_id"], 0.0)
        value += (
            flow
            * conversion_rate(store["store_type"], hour)
            * store["price"]
            * store["attraction"]
            * time_multiplier(store["store_type"], hour)
            * flow_bonus
            * area_fit
            * anchor_spillover
            * category_bonus
        )
    return value


def precompute_score_table(
    stores: dict[str, dict[str, Any]],
    positions: dict[str, dict[str, Any]],
    node_flow: dict[int, dict[str, float]],
) -> dict[str, dict[str, float]]:
    table: dict[str, dict[str, float]] = {}
    for position_id, position in positions.items():
        table[position_id] = {}
        for store_id, store in stores.items():
            table[position_id][store_id] = position_score(store, position, node_flow)
    return table


def layout_revenue(layout: dict[str, str], score_table: dict[str, dict[str, float]]) -> float:
    return sum(score_table[position_id][store_id] for position_id, store_id in layout.items())


def try_swap(
    layout: dict[str, str],
    p1: str,
    p2: str,
    stores: dict[str, dict[str, Any]],
    positions: dict[str, dict[str, Any]],
) -> dict[str, str] | None:
    s1, s2 = layout[p1], layout[p2]
    if not is_assignment_feasible(stores[s1], positions[p2]):
        return None
    if not is_assignment_feasible(stores[s2], positions[p1]):
        return None
    new_layout = dict(layout)
    new_layout[p1], new_layout[p2] = s2, s1
    return new_layout


def optimize_layout(
    stores: dict[str, dict[str, Any]],
    positions: dict[str, dict[str, Any]],
    real_layout: dict[str, str],
    score_table: dict[str, dict[str, float]],
) -> tuple[dict[str, str], list[dict[str, Any]]]:
    position_ids = list(real_layout)
    trace: list[dict[str, Any]] = []
    global_best = dict(real_layout)
    global_best_score = layout_revenue(global_best, score_table)
    real_score = global_best_score

    seeds = [RANDOM_SEED + offset for offset in range(8)]
    iterations = 1600
    cooling = 0.995

    for seed_index, seed in enumerate(seeds, start=1):
        random.seed(seed)
        current = dict(real_layout)
        for _ in range(20 * seed_index):
            p1, p2 = random.sample(position_ids, 2)
            candidate = try_swap(current, p1, p2, stores, positions)
            if candidate is not None:
                current = candidate
        current_score = layout_revenue(current, score_table)
        best = dict(current)
        best_score = current_score
        t0 = max(real_score * 0.020, 1.0)

        for it in range(1, iterations + 1):
            temperature = max(1.0, t0 * (cooling ** it))
            p1, p2 = random.sample(position_ids, 2)
            candidate = try_swap(current, p1, p2, stores, positions)
            accepted = False
            if candidate is not None:
                new_score = layout_revenue(candidate, score_table)
                diff = new_score - current_score
                if diff >= 0 or random.random() < math.exp(diff / temperature):
                    current = candidate
                    current_score = new_score
                    accepted = True
                    if new_score > best_score:
                        best = dict(candidate)
                        best_score = new_score
                    if new_score > global_best_score:
                        global_best = dict(candidate)
                        global_best_score = new_score

            trace.append(
                {
                    "seed": seed_index,
                    "iteration": it,
                    "current_score": current_score,
                    "best_score": best_score,
                    "global_best_score": global_best_score,
                    "temperature": temperature,
                    "accepted": 1 if accepted else 0,
                }
            )

        improved = True
        local_round = 0
        while improved and local_round < 80:
            improved = False
            local_round += 1
            p1, p2 = random.sample(position_ids, 2)
            candidate = try_swap(global_best, p1, p2, stores, positions)
            if candidate is not None:
                score = layout_revenue(candidate, score_table)
                if score > global_best_score:
                    global_best = candidate
                    global_best_score = score
                    improved = True

    return global_best, trace


def inverse_layout(layout: dict[str, str]) -> dict[str, str]:
    return {store_id: position_id for position_id, store_id in layout.items()}


def build_comparison_rows(
    stores: dict[str, dict[str, Any]],
    positions: dict[str, dict[str, Any]],
    real_layout: dict[str, str],
    best_layout: dict[str, str],
    score_table: dict[str, dict[str, float]],
) -> list[dict[str, Any]]:
    real_pos_by_store = inverse_layout(real_layout)
    best_pos_by_store = inverse_layout(best_layout)
    rows = []
    for store_id, store in stores.items():
        real_pos = real_pos_by_store[store_id]
        opt_pos = best_pos_by_store[store_id]
        real_revenue = score_table[real_pos][store_id]
        opt_revenue = score_table[opt_pos][store_id]
        gain = opt_revenue - real_revenue
        gain_rate = gain / real_revenue if real_revenue > 0 else 0
        rows.append(
            {
                "store_id": store_id,
                "brand": store["brand"],
                "store_type": store["store_type"],
                "category_text": store["category_text"],
                "real_position": real_pos,
                "optimized_position": opt_pos,
                "real_node": positions[real_pos]["node_id"],
                "optimized_node": positions[opt_pos]["node_id"],
                "real_floor": positions[real_pos]["floor"],
                "optimized_floor": positions[opt_pos]["floor"],
                "real_revenue": real_revenue,
                "optimized_revenue": opt_revenue,
                "gain": gain,
                "gain_rate": gain_rate,
                "moved": 1 if real_pos != opt_pos else 0,
            }
        )
    return rows


def write_csv_outputs(rows: list[dict[str, Any]], trace: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    with COMPARISON_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        fieldnames = [
            "store_id",
            "brand",
            "store_type",
            "category_text",
            "real_position",
            "optimized_position",
            "real_node",
            "optimized_node",
            "real_floor",
            "optimized_floor",
            "real_revenue",
            "optimized_revenue",
            "gain",
            "gain_rate",
            "moved",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            out = dict(row)
            for key in ["real_revenue", "optimized_revenue", "gain", "gain_rate"]:
                out[key] = round(out[key], 4)
            writer.writerow(out)

    with TRACE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        fieldnames = ["seed", "iteration", "current_score", "best_score", "global_best_score", "temperature", "accepted"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in trace:
            out = dict(row)
            out["current_score"] = round(out["current_score"], 4)
            out["best_score"] = round(out["best_score"], 4)
            out["global_best_score"] = round(out["global_best_score"], 4)
            out["temperature"] = round(out["temperature"], 4)
            writer.writerow(out)

    with SUMMARY_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["metric", "value"])
        for key, value in summary.items():
            writer.writerow([key, value])



def draw_store_gain_top10(rows: list[dict[str, Any]]) -> None:
    setup_matplotlib_style()
    top_rows = sorted(rows, key=lambda r: r["gain"], reverse=True)[:10]
    fig, ax = plt.subplots(figsize=(10.5, 6.2))
    labels = [short_label(row["brand"], 12) for row in top_rows]
    values = [row["gain"] / 10000 for row in top_rows]
    ax.barh(range(len(labels)), values, color=[PURPLE_DARK] + [BLUE_DARK] * (len(labels) - 1), label="收益提升")
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels)
    ax.invert_yaxis()
    ax.set_xlabel("万元")
    place_legend(ax)
    save_figure(fig, FIG_GAIN_TOP10)


def draw_task2_dashboards(rows: list[dict[str, Any]], trace: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    setup_matplotlib_style()
    fig, axes = plt.subplots(2, 2, figsize=(15.5, 9.2))
    ax = axes[0, 0]
    first_seed = min(int(row["seed"]) for row in trace)
    first_trace = [row for row in trace if int(row["seed"]) == first_seed]
    xs = [row["iteration"] for row in first_trace]
    ax.plot(xs, [row["current_score"] / 10000 for row in first_trace], color=BLUE_MID, label="当前方案")
    ax.plot(xs, [row["best_score"] / 10000 for row in first_trace], color=PURPLE_DARK, label="该起点最优")
    ax.set_title("(a) 单起点收敛曲线")
    ax.set_xlabel("迭代次数")
    ax.set_ylabel("估计营业额（万元）")
    place_legend(ax)

    ax = axes[0, 1]
    by_seed: dict[int, list[float]] = defaultdict(list)
    for row in trace:
        by_seed[int(row["seed"])].append(row["best_score"] / 10000)
    colors = [BLUE_DARK, GREEN_DARK, PURPLE_DARK, BLUE_MID, GREEN_MID, PURPLE_MID]
    for idx, (seed, values) in enumerate(sorted(by_seed.items())):
        step = max(1, len(values) // 300)
        sampled = values[::step]
        ax.plot(range(len(sampled)), sampled, color=colors[idx % len(colors)], linewidth=1.5, alpha=0.9, label=f"起点{seed}")
    ax.set_title("(b) 多起点收敛对比")
    ax.set_xlabel("迭代进度")
    ax.set_ylabel("历史最优（万元）")
    place_legend(ax, ncol=2)

    ax = axes[1, 0]
    values = [float(summary["real_total_revenue"]) / 10000, float(summary["optimized_total_revenue"]) / 10000]
    ax.bar([0, 1], values, color=[BLUE_MID, PURPLE_DARK], label="总营业额")
    ax.set_xticks([0, 1])
    ax.set_xticklabels(["现实布局", "优化布局"])
    ax.set_ylabel("万元")
    ax.set_title("(c) 优化前后总营业额")
    place_legend(ax)

    ax = axes[1, 1]
    gains = [row["gain_rate"] * 100 for row in rows]
    bins = [-40, -20, 0, 10, 20, 40, 80, 160]
    ax.hist(gains, bins=bins, color=GREEN_DARK, edgecolor="white", label="店铺数量")
    ax.set_title("(d) 收益提升率分布")
    ax.set_xlabel("提升率（%）")
    ax.set_ylabel("数量")
    place_legend(ax)
    save_figure(fig, FIG_OPT_DASHBOARD)

    fig, axes = plt.subplots(2, 2, figsize=(15.5, 9.5))
    ax = axes[0, 0]
    matrix = [[0 for _ in FLOORS] for _ in FLOORS]
    floor_index = {floor: idx for idx, floor in enumerate(FLOORS)}
    for row in rows:
        if row["real_floor"] in floor_index and row["optimized_floor"] in floor_index:
            matrix[floor_index[row["real_floor"]]][floor_index[row["optimized_floor"]]] += 1
    im = ax.imshow(matrix, cmap=cmap_from(PURPLE_LIGHT, PURPLE_DARK, "migration"))
    ax.set_title("(a) 楼层迁移矩阵")
    ax.set_xticks(range(len(FLOORS)))
    ax.set_xticklabels(FLOORS)
    ax.set_yticks(range(len(FLOORS)))
    ax.set_yticklabels(FLOORS)
    ax.set_xlabel("优化后楼层")
    ax.set_ylabel("原楼层")
    for i in range(len(FLOORS)):
        for j in range(len(FLOORS)):
            ax.text(j, i, str(matrix[i][j]), ha="center", va="center", color=TEXT_DARK)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    ax = axes[0, 1]
    real = Counter()
    opt = Counter()
    for row in rows:
        real[row["real_floor"]] += row["real_revenue"]
        opt[row["optimized_floor"]] += row["optimized_revenue"]
    heat = [[real[floor] / 10000 for floor in FLOORS], [opt[floor] / 10000 for floor in FLOORS]]
    im = ax.imshow(heat, aspect="auto", cmap=cmap_from(GREEN_LIGHT, GREEN_DARK, "floor_rev"))
    ax.set_title("(b) 优化前后楼层收益")
    ax.set_xticks(range(len(FLOORS)))
    ax.set_xticklabels(FLOORS)
    ax.set_yticks([0, 1])
    ax.set_yticklabels(["现实布局", "优化布局"])
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="万元")

    ax = axes[1, 0]
    type_matrix = []
    type_labels = [("dining", "餐饮店"), ("retail", "零售店"), ("anchor", "主力店")]
    for typ, _label in type_labels:
        counter = Counter(row["optimized_floor"] for row in rows if row["store_type"] == typ)
        type_matrix.append([counter[floor] for floor in FLOORS])
    im = ax.imshow(type_matrix, aspect="auto", cmap=cmap_from(BLUE_LIGHT, BLUE_DARK, "type_floor"))
    ax.set_title("(c) 优化后业态楼层分布")
    ax.set_xticks(range(len(FLOORS)))
    ax.set_xticklabels(FLOORS)
    ax.set_yticks(range(len(type_labels)))
    ax.set_yticklabels([label for _typ, label in type_labels])
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    ax = axes[1, 1]
    values = [
        sum(1 for r in rows if r["gain"] > 0),
        sum(1 for r in rows if not r["moved"]),
        sum(1 for r in rows if r["moved"]),
        sum(1 for r in rows if r["gain"] < 0),
    ]
    labels = ["收益提升", "位置保留", "发生迁移", "收益下降"]
    ax.bar(labels, values, color=[BLUE_DARK, GREEN_DARK, PURPLE_DARK, PURPLE_MID], label="店铺数量")
    ax.set_title("(d) 布局约束与结果诊断")
    ax.set_ylabel("数量")
    place_legend(ax)
    save_figure(fig, FIG_LAYOUT_DASHBOARD)
    draw_store_gain_top10(rows)


def write_all_figures(rows: list[dict[str, Any]], trace: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    draw_task2_dashboards(rows, trace, summary)

def build_summary(rows: list[dict[str, Any]], real_score: float, opt_score: float, trace: list[dict[str, Any]]) -> dict[str, Any]:
    moved = sum(row["moved"] for row in rows)
    summary = {
        "store_count": len(rows),
        "real_total_revenue": round(real_score, 4),
        "optimized_total_revenue": round(opt_score, 4),
        "revenue_gain": round(opt_score - real_score, 4),
        "improvement_rate": round((opt_score - real_score) / real_score, 6),
        "moved_store_count": moved,
        "move_rate": round(moved / len(rows), 6),
        "iterations": len(trace),
        "accepted_steps": sum(row["accepted"] for row in trace),
    }
    return summary


def main() -> None:
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    stores, positions, real_layout = load_stores_and_positions()
    node_flow = load_node_flow(positions)
    score_table = precompute_score_table(stores, positions, node_flow)
    real_score = layout_revenue(real_layout, score_table)
    best_layout, trace = optimize_layout(stores, positions, real_layout, score_table)
    opt_score = layout_revenue(best_layout, score_table)
    rows = build_comparison_rows(stores, positions, real_layout, best_layout, score_table)
    summary = build_summary(rows, real_score, opt_score, trace)

    write_csv_outputs(rows, trace, summary)
    write_all_figures(rows, trace, summary)

    print("Task 2 layout optimization finished.")
    print(f"Stores: {len(stores)}")
    print(f"Positions: {len(positions)}")
    print(f"Real revenue: {real_score:.2f}")
    print(f"Optimized revenue: {opt_score:.2f}")
    print(f"Improvement: {(opt_score - real_score) / real_score * 100:.2f}%")
    print(f"Moved stores: {summary['moved_store_count']}")
    print(f"Wrote: {COMPARISON_CSV}")
    print(f"Wrote: {TRACE_CSV}")
    print(f"Wrote: {SUMMARY_CSV}")
if __name__ == "__main__":
    main()
