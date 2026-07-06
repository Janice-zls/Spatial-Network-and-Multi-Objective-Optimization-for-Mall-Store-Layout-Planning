

import csv
import math
import random
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import matplotlib.pyplot as plt

from plot_style import (
    BLUE_DARK,
    BLUE_LIGHT,
    BLUE_MID,
    COLOR_WHITE,
    GREEN_DARK,
    GREEN_LIGHT,
    GREEN_MID,
    PURPLE_DARK,
    PURPLE_LIGHT,
    PURPLE_MID,
    TEXT_DARK,
    cmap_from,
    place_legend,
    save_figure,
    short_label,
    setup_matplotlib_style,
)

ROOT = Path("SJMMA/Problem_B")
PROJECT_DIR = Path("SJMMA/2026_SJMMA")
FIGURE_DIR = PROJECT_DIR / "figures"

COMPARISON_CSV = ROOT / "task2_layout_comparison.csv"
NODE_FLOW_CSV = ROOT / "task1_node_flow_by_hour.csv"
WORKBOOK_PATH = ROOT / "guangzhou_taikoo_hui_store_directory_integrated.xlsx"

CANDIDATE_CSV = ROOT / "task3_candidate_solutions.csv"
STORE_DETAIL_CSV = ROOT / "task3_store_profit_fairness.csv"
SUMMARY_CSV = ROOT / "task3_summary.csv"

FIG_PARETO = FIGURE_DIR / "task3_pareto_revenue_satisfaction.png"
FIG_PARETO_PROFIT_SATISFACTION = FIGURE_DIR / "task3_pareto_profit_satisfaction.png"
FIG_PARETO_REVENUE_PROFIT = FIGURE_DIR / "task3_pareto_revenue_profit.png"
FIG_PARETO_REVENUE_FAIRNESS = FIGURE_DIR / "task3_pareto_revenue_fairness.png"
FIG_PARETO_PROFIT_FAIRNESS = FIGURE_DIR / "task3_pareto_profit_fairness.png"
FIG_PARETO_FAIRNESS_SATISFACTION = FIGURE_DIR / "task3_pareto_fairness_satisfaction.png"
FIG_RADAR = FIGURE_DIR / "task3_objective_radar.png"
FIG_GINI = FIGURE_DIR / "task3_gini_by_solution.png"
FIG_PROFIT_REVENUE = FIGURE_DIR / "task3_profit_vs_revenue.png"
FIG_SATISFACTION = FIGURE_DIR / "task3_satisfaction_components.png"
FIG_PARETO_MATRIX = FIGURE_DIR / "task3_pareto_2d_matrix.png"
FIG_PARALLEL = FIGURE_DIR / "task3_parallel_coordinates.png"
FIG_TRADEOFF = FIGURE_DIR / "task3_objective_tradeoff_heatmap.png"
FIG_SCORE_TABLE = FIGURE_DIR / "task3_solution_score_table.png"
FIG_FAIRNESS_PROFIT = FIGURE_DIR / "task3_fairness_profit_scatter.png"
FIG_MULTI_DASHBOARD = FIGURE_DIR / "task3_multiobjective_dashboard.png"
FIG_PARETO_DASHBOARD = FIGURE_DIR / "task3_pareto_diagnostics_dashboard.png"

RANDOM_SEED = 20260502

FLOOR_RENT_FACTOR = {"M": 1.25, "MU": 1.15, "L1": 1.35, "L2": 1.05, "L3": 0.90}
FLOOR_ACCESS = {"M": 0.86, "MU": 0.90, "L1": 1.00, "L2": 0.82, "L3": 0.70}
STORE_TYPES = ["dining", "retail", "anchor"]
TYPE_CN = {"dining": "餐饮店", "retail": "零售店", "anchor": "主力店"}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_positions() -> dict[str, dict[str, Any]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["整合主表"]
    headers = [c.value for c in ws[1]]
    counter: Counter[str] = Counter()
    positions = {}
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, row_values))
        brand = clean(row.get("店铺/品牌"))
        node = clean(row.get("node_id"))
        if not brand or not node:
            continue
        counter[node] += 1
        position_id = f"{node}@{counter[node]}"
        positions[position_id] = {
            "position_id": position_id,
            "node_id": node,
            "floor": clean(row.get("楼层")),
            "node_category": clean(row.get("节点类别")),
            "size_rating": int(row.get("size_rating") or 1),
        }
    return positions


def read_node_heat() -> dict[str, float]:
    heat: Counter[str] = Counter()
    if NODE_FLOW_CSV.exists():
        with NODE_FLOW_CSV.open("r", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                heat[row["node_id"]] += float(row["flow"])
    max_heat = max(heat.values()) if heat else 1.0
    return {node: value / max_heat for node, value in heat.items()}


def read_task2_rows() -> list[dict[str, Any]]:
    rows = []
    with COMPARISON_CSV.open("r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            item = dict(row)
            for key in ["real_revenue", "optimized_revenue", "gain", "gain_rate"]:
                item[key] = float(item[key])
            item["moved"] = int(item["moved"])
            rows.append(item)
    return rows


def margin_rate(store_type: str) -> float:
    return {"dining": 0.42, "retail": 0.34, "anchor": 0.24}.get(store_type, 0.32)


def operating_cost(store_type: str) -> float:
    return {"dining": 18000.0, "retail": 26000.0, "anchor": 42000.0}.get(store_type, 22000.0)


def rent_for_position(position_id: str, positions: dict[str, dict[str, Any]], heat: dict[str, float]) -> float:
    pos = positions[position_id]
    floor_factor = FLOOR_RENT_FACTOR.get(pos["floor"], 1.0)
    heat_factor = 0.65 + 0.90 * heat.get(pos["node_id"], 0.2)
    return 14500.0 * floor_factor * pos["size_rating"] * heat_factor


def gini(values: list[float]) -> float:
    values = [max(0.0, v) for v in values]
    total = sum(values)
    n = len(values)
    if n == 0 or total == 0:
        return 0.0
    diff_sum = 0.0
    for a in values:
        for b in values:
            diff_sum += abs(a - b)
    return diff_sum / (2 * n * total)


def solution_layout(rows: list[dict[str, Any]], mode: str) -> dict[str, str]:
    layout = {}
    for row in rows:
        if mode == "real":
            layout[row["store_id"]] = row["real_position"]
        else:
            layout[row["store_id"]] = row["optimized_position"]
    return layout


def make_candidate_layouts(rows: list[dict[str, Any]], positions: dict[str, dict[str, Any]]) -> dict[str, dict[str, str]]:
    random.seed(RANDOM_SEED)
    candidates = {
        "现实布局": solution_layout(rows, "real"),
        "营业额最大方案": solution_layout(rows, "optimized"),
    }

    opt = candidates["营业额最大方案"]
    rent_layout = dict(opt)
    for row in rows:
        pos = opt[row["store_id"]]
        if positions[pos]["floor"] == "L1" and row["gain_rate"] < 0.12:
            rent_layout[row["store_id"]] = row["real_position"]
    candidates["租金利润方案"] = rent_layout
    fair_layout = dict(opt)
    for row in rows:
        if row["gain_rate"] < 0.05:
            fair_layout[row["store_id"]] = row["real_position"]
    candidates["公平方案"] = fair_layout
    sat_layout = dict(opt)
    for row in rows:
        if row["store_type"] in {"dining", "anchor"} and random.random() < 0.55:
            sat_layout[row["store_id"]] = row["real_position"]
    candidates["满意度方案"] = sat_layout
    base_ids = [row["store_id"] for row in rows]
    for idx in range(1, 49):
        layout = dict(opt)
        for _ in range(8 + idx // 2):
            a, b = random.sample(base_ids, 2)
            layout[a], layout[b] = layout[b], layout[a]
        candidates[f"进化扰动方案{idx}"] = layout

    return candidates


def revenue_for_store(row: dict[str, Any], position_id: str) -> float:
    if position_id == row["real_position"]:
        return row["real_revenue"]
    if position_id == row["optimized_position"]:
        return row["optimized_revenue"]
    base = 0.5 * row["real_revenue"] + 0.5 * row["optimized_revenue"]
    return base * random.uniform(0.88, 1.08)


def evaluate_solution(
    name: str,
    layout: dict[str, str],
    rows: list[dict[str, Any]],
    positions: dict[str, dict[str, Any]],
    heat: dict[str, float],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    details = []
    revenues = []
    profits = []
    growth_rates = []
    rent_total = 0.0
    access_scores = []
    crowd_penalties = []
    walking_penalties = []
    type_floor = {typ: set() for typ in STORE_TYPES}

    for row in rows:
        position_id = layout[row["store_id"]]
        pos = positions[position_id]
        revenue = revenue_for_store(row, position_id)
        rent = rent_for_position(position_id, positions, heat)
        profit = margin_rate(row["store_type"]) * revenue - rent - operating_cost(row["store_type"])
        growth = (revenue - row["real_revenue"]) / row["real_revenue"] if row["real_revenue"] > 0 else 0.0
        floor_access = FLOOR_ACCESS.get(pos["floor"], 0.75)
        node_heat = heat.get(pos["node_id"], 0.2)
        crowd_penalty = max(0.0, node_heat - 0.70)
        walking_penalty = 0.08 if pos["floor"] != row["real_floor"] and row["store_type"] in {"anchor", "dining"} else 0.02

        revenues.append(revenue)
        profits.append(profit)
        growth_rates.append(growth)
        rent_total += rent
        access_scores.append(floor_access)
        crowd_penalties.append(crowd_penalty)
        walking_penalties.append(walking_penalty)
        type_floor[row["store_type"]].add(pos["floor"])
        details.append({
            "solution": name,
            "store_id": row["store_id"],
            "brand": row["brand"],
            "store_type": row["store_type"],
            "position_id": position_id,
            "floor": pos["floor"],
            "revenue": revenue,
            "rent": rent,
            "profit": profit,
            "growth_rate": growth,
        })

    revenue_total = sum(revenues)
    profit_total = sum(profits)
    gini_value = gini(revenues)
    fair_score = 1.0 - gini_value
    access = sum(access_scores) / len(access_scores)
    variety = sum(len(type_floor[typ]) for typ in STORE_TYPES) / (len(STORE_TYPES) * 5.0)
    crowd = sum(crowd_penalties) / len(crowd_penalties)
    walking = sum(walking_penalties) / len(walking_penalties)
    satisfaction = 0.40 * access + 0.30 * variety - 0.18 * crowd - 0.12 * walking
    metric = {
        "solution": name,
        "revenue": revenue_total,
        "profit": profit_total,
        "rent": rent_total,
        "gini": gini_value,
        "fairness": fair_score,
        "satisfaction": satisfaction,
        "access_score": access,
        "variety_score": variety,
        "crowd_penalty": crowd,
        "walking_penalty": walking,
    }
    return metric, details


def normalize_metrics(metrics: list[dict[str, Any]]) -> None:
    keys = ["revenue", "profit", "fairness", "satisfaction"]
    for key in keys:
        vals = [m[key] for m in metrics]
        lo, hi = min(vals), max(vals)
        for m in metrics:
            m[f"{key}_norm"] = 1.0 if hi == lo else (m[key] - lo) / (hi - lo)
    for m in metrics:
        m["combined_score"] = (
            0.32 * m["revenue_norm"]
            + 0.28 * m["profit_norm"]
            + 0.20 * m["fairness_norm"]
            + 0.20 * m["satisfaction_norm"]
        )


def dominates(a: dict[str, Any], b: dict[str, Any]) -> bool:
    keys = ["revenue", "profit", "fairness", "satisfaction"]
    return all(a[k] >= b[k] for k in keys) and any(a[k] > b[k] for k in keys)


def assign_nsga_rank_and_crowding(metrics: list[dict[str, Any]]) -> None:
    remaining = set(range(len(metrics)))
    rank = 1
    while remaining:
        front = []
        for idx in list(remaining):
            if not any(j != idx and j in remaining and dominates(metrics[j], metrics[idx]) for j in remaining):
                front.append(idx)
        for idx in front:
            metrics[idx]["nsga_rank"] = rank
            remaining.remove(idx)
        rank += 1

    keys = ["revenue_norm", "profit_norm", "fairness_norm", "satisfaction_norm"]
    for m in metrics:
        m["crowding_distance"] = 0.0
    fronts = defaultdict(list)
    for idx, m in enumerate(metrics):
        fronts[m["nsga_rank"]].append(idx)
    for idxs in fronts.values():
        if len(idxs) <= 2:
            for idx in idxs:
                metrics[idx]["crowding_distance"] = 999.0
            continue
        for key in keys:
            ordered = sorted(idxs, key=lambda idx: metrics[idx][key])
            metrics[ordered[0]]["crowding_distance"] = 999.0
            metrics[ordered[-1]]["crowding_distance"] = 999.0
            lo = metrics[ordered[0]][key]
            hi = metrics[ordered[-1]][key]
            span = hi - lo or 1.0
            for pos in range(1, len(ordered) - 1):
                prev_v = metrics[ordered[pos - 1]][key]
                next_v = metrics[ordered[pos + 1]][key]
                metrics[ordered[pos]]["crowding_distance"] += (next_v - prev_v) / span


def mark_pareto(metrics: list[dict[str, Any]]) -> None:
    for m in metrics:
        m["is_pareto"] = 1
        for other in metrics:
            if other is not m and dominates(other, m):
                m["is_pareto"] = 0
                break


def write_csv_outputs(metrics: list[dict[str, Any]], details: list[dict[str, Any]]) -> dict[str, Any]:
    with CANDIDATE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        keys = [
            "solution", "revenue", "profit", "rent", "gini", "fairness", "satisfaction",
            "access_score", "variety_score", "crowd_penalty", "walking_penalty",
            "combined_score", "nsga_rank", "crowding_distance", "is_pareto",
        ]
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        for m in metrics:
            writer.writerow({k: round(m[k], 6) if isinstance(m[k], float) else m[k] for k in keys})

    with STORE_DETAIL_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        keys = ["solution", "store_id", "brand", "store_type", "position_id", "floor", "revenue", "rent", "profit", "growth_rate"]
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        for d in details:
            writer.writerow({k: round(d[k], 6) if isinstance(d[k], float) else d[k] for k in keys})

    best_compromise = max(metrics, key=lambda m: m["combined_score"])
    best_revenue = max(metrics, key=lambda m: m["revenue"])
    best_satisfaction = max(metrics, key=lambda m: m["satisfaction"])
    summary = {
        "candidate_count": len(metrics),
        "pareto_count": sum(m["is_pareto"] for m in metrics),
        "best_compromise_solution": best_compromise["solution"],
        "best_revenue_solution": best_revenue["solution"],
        "best_satisfaction_solution": best_satisfaction["solution"],
        "max_revenue": round(best_revenue["revenue"], 4),
        "max_profit": round(max(m["profit"] for m in metrics), 4),
        "min_gini": round(min(m["gini"] for m in metrics), 6),
        "max_satisfaction": round(best_satisfaction["satisfaction"], 6),
    }
    with SUMMARY_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["metric", "value"])
        for k, v in summary.items():
            writer.writerow([k, v])
    return summary



def main_solutions(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [m for m in metrics if not m["solution"].startswith("进化扰动")][:6]


def draw_solution_score_table(metrics: list[dict[str, Any]]) -> None:
    setup_matplotlib_style()
    main = sorted(metrics, key=lambda m: (m["nsga_rank"], -m["combined_score"]))[:8]
    fig, ax = plt.subplots(figsize=(12.5, 4.6))
    ax.axis("off")
    headers = ["方案", "营业额", "净利润", "公平性", "满意度", "综合分", "Rank"]
    table_data = []
    for m in main:
        table_data.append([
            short_label(m["solution"], 10),
            f"{m['revenue_norm']:.2f}",
            f"{m['profit_norm']:.2f}",
            f"{m['fairness_norm']:.2f}",
            f"{m['satisfaction_norm']:.2f}",
            f"{m['combined_score']:.2f}",
            str(m["nsga_rank"]),
        ])
    table = ax.table(cellText=table_data, colLabels=headers, cellLoc="center", loc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 1.55)
    for (row, _col), cell in table.get_celld().items():
        if row == 0:
            cell.set_facecolor(BLUE_LIGHT)
            cell.set_text_props(weight="bold")
        elif row % 2 == 0:
            cell.set_facecolor(GREEN_LIGHT)
        else:
            cell.set_facecolor(PURPLE_LIGHT)
    ax.set_title("代表方案综合评分表", fontsize=16, fontweight="bold", pad=18)
    save_figure(fig, FIG_SCORE_TABLE)


def draw_objective_radar(metrics: list[dict[str, Any]]) -> None:
    setup_matplotlib_style()
    chosen = [m for m in metrics if m["solution"] in {"营业额最大方案", "公平方案", "满意度方案", "折中方案"}][:4]
    labels = ["营业额", "净利润", "公平性", "满意度"]
    angles = [i / len(labels) * 2 * math.pi for i in range(len(labels))]
    angles += angles[:1]

    fig = plt.figure(figsize=(17, 15))
    ax = fig.add_subplot(111, projection="polar")
    fig.subplots_adjust(left=0.07, right=0.76, top=0.88, bottom=0.08)
    for m, color in zip(chosen, [BLUE_DARK, GREEN_DARK, PURPLE_DARK, PURPLE_MID]):
        values = [m["revenue_norm"], m["profit_norm"], m["fairness_norm"], m["satisfaction_norm"]]
        values += values[:1]
        ax.plot(angles, values, color=color, linewidth=3, label=m["solution"])
        ax.fill(angles, values, color=color, alpha=0.12)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=24, fontweight="bold")
    ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
    ax.set_yticklabels(["0.2", "0.4", "0.6", "0.8", "1.0"], fontsize=17, fontweight="bold")
    ax.set_ylim(0, 1)
    ax.set_title("四目标雷达图", fontsize=34, fontweight="bold", pad=38)
    legend = ax.legend(
        loc="upper right",
        bbox_to_anchor=(0.89, 0.93),
        bbox_transform=fig.transFigure,
        framealpha=0.95,
        fontsize=19,
        borderaxespad=0.0,
    )
    for text in legend.get_texts():
        text.set_fontweight("bold")
    FIG_RADAR.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(FIG_RADAR, dpi=220, facecolor=COLOR_WHITE)
    plt.close(fig)


def draw_pareto_pair(
    metrics: list[dict[str, Any]],
    x_key: str,
    y_key: str,
    x_label: str,
    y_label: str,
    title: str,
    path: Path,
    x_scale: float = 1.0,
    y_scale: float = 1.0,
    legend_inside: bool = False,
) -> None:
    setup_matplotlib_style()
    ordinary = [m for m in metrics if not m["is_pareto"]]
    pareto = sorted([m for m in metrics if m["is_pareto"]], key=lambda m: m[x_key])

    fig, ax = plt.subplots(figsize=(17, 15))
    fig.subplots_adjust(left=0.10, right=0.78, top=0.88, bottom=0.10)
    ax.scatter(
        [m[x_key] / x_scale for m in ordinary],
        [m[y_key] / y_scale for m in ordinary],
        color=PURPLE_LIGHT,
        edgecolor=PURPLE_MID,
        linewidth=0.8,
        s=110,
        alpha=0.80,
        label="普通方案",
    )
    ax.scatter(
        [m[x_key] / x_scale for m in pareto],
        [m[y_key] / y_scale for m in pareto],
        color=PURPLE_DARK,
        edgecolor=COLOR_WHITE,
        linewidth=1.3,
        s=190,
        alpha=0.96,
        label="帕累托方案",
        zorder=3,
    )
    if pareto:
        ax.plot(
            [m[x_key] / x_scale for m in pareto],
            [m[y_key] / y_scale for m in pareto],
            color=PURPLE_MID,
            linewidth=2.8,
            alpha=0.85,
            label="帕累托前沿",
            zorder=2,
        )

    ax.set_title(title, fontsize=34, fontweight="bold", pad=26)
    ax.set_xlabel(x_label, fontsize=24, fontweight="bold")
    ax.set_ylabel(y_label, fontsize=24, fontweight="bold")
    ax.tick_params(axis="both", labelsize=19)
    for tick in ax.get_xticklabels() + ax.get_yticklabels():
        tick.set_fontweight("bold")
    if legend_inside:
        legend = ax.legend(loc="upper right", framealpha=0.95, fontsize=19)
    else:
        legend = ax.legend(
            loc="upper right",
            bbox_to_anchor=(0.97, 0.94),
            bbox_transform=fig.transFigure,
            framealpha=0.95,
            fontsize=19,
            borderaxespad=0.0,
        )
    for text in legend.get_texts():
        text.set_fontweight("bold")
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=220, facecolor=COLOR_WHITE)
    plt.close(fig)


def draw_pareto_revenue_satisfaction(metrics: list[dict[str, Any]]) -> None:
    draw_pareto_pair(
        metrics,
        "revenue",
        "satisfaction",
        "营业额（万元）",
        "满意度",
        "营业额-满意度帕累托前沿",
        FIG_PARETO,
        x_scale=10000,
    )


def draw_pareto_pairwise_choices(metrics: list[dict[str, Any]]) -> None:
    pair_specs = [
        ("profit", "satisfaction", "净利润（万元）", "满意度", "净利润-满意度帕累托前沿", FIG_PARETO_PROFIT_SATISFACTION, 10000, 1),
        ("revenue", "profit", "营业额（万元）", "净利润（万元）", "营业额-净利润帕累托前沿", FIG_PARETO_REVENUE_PROFIT, 10000, 10000),
        ("revenue", "fairness", "营业额（万元）", "公平性", "营业额-公平性帕累托前沿", FIG_PARETO_REVENUE_FAIRNESS, 10000, 1),
        ("profit", "fairness", "净利润（万元）", "公平性", "净利润-公平性帕累托前沿", FIG_PARETO_PROFIT_FAIRNESS, 10000, 1, True),
        ("fairness", "satisfaction", "公平性", "满意度", "公平性-满意度帕累托前沿", FIG_PARETO_FAIRNESS_SATISFACTION, 1, 1),
    ]
    for spec in pair_specs:
        x_key, y_key, x_label, y_label, title, path, x_scale, y_scale, *options = spec
        draw_pareto_pair(
            metrics,
            x_key,
            y_key,
            x_label,
            y_label,
            title,
            path,
            x_scale=x_scale,
            y_scale=y_scale,
            legend_inside=bool(options and options[0]),
        )


def draw_task3_dashboards(metrics: list[dict[str, Any]]) -> None:
    setup_matplotlib_style()
    main = main_solutions(metrics)

    fig = plt.figure(figsize=(15.5, 9.2))
    gs = fig.add_gridspec(2, 2)

    ax = fig.add_subplot(gs[0, 0])
    revs = [m["revenue"] / 10000 for m in metrics]
    sats = [m["satisfaction"] for m in metrics]
    colors = [PURPLE_DARK if m["is_pareto"] else BLUE_MID for m in metrics]
    ax.scatter(revs, sats, c=colors, s=[55 if m["is_pareto"] else 25 for m in metrics], label="候选方案")
    ax.set_title("(a) 营业额-满意度帕累托散点")
    ax.set_xlabel("营业额（万元）")
    ax.set_ylabel("满意度")
    ax.scatter([], [], color=PURPLE_DARK, label="帕累托方案")
    ax.scatter([], [], color=BLUE_MID, label="普通方案")
    place_legend(ax)

    ax = fig.add_subplot(gs[0, 1], projection="polar")
    chosen = [m for m in metrics if m["solution"] in {"营业额最大方案", "公平方案", "满意度方案", "折中方案"}][:4]
    labels = ["营业额", "净利润", "公平性", "满意度"]
    angles = [i / len(labels) * 2 * math.pi for i in range(len(labels))]
    angles += angles[:1]
    for m, color in zip(chosen, [BLUE_DARK, GREEN_DARK, PURPLE_DARK, BLUE_MID]):
        values = [m["revenue_norm"], m["profit_norm"], m["fairness_norm"], m["satisfaction_norm"]]
        values += values[:1]
        ax.plot(angles, values, color=color, linewidth=2, label=m["solution"])
        ax.fill(angles, values, color=color, alpha=0.12)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels)
    ax.set_ylim(0, 1)
    ax.set_title("(b) 四目标雷达图", y=1.08)
    place_legend(ax)

    ax = fig.add_subplot(gs[1, 0])
    ax.bar([m["solution"] for m in main], [m["gini"] for m in main], color=[BLUE_DARK, GREEN_DARK, PURPLE_DARK, BLUE_MID, GREEN_MID, PURPLE_MID][: len(main)], label="Gini系数")
    ax.set_title("(c) 不同方案 Gini 系数")
    ax.set_ylabel("Gini")
    ax.tick_params(axis="x", rotation=20)
    place_legend(ax)

    ax = fig.add_subplot(gs[1, 1])
    xs = list(range(len(main)))
    width = 0.35
    ax.bar([i - width / 2 for i in xs], [m["revenue"] / 10000 for m in main], width=width, color=BLUE_DARK, label="营业额")
    ax.bar([i + width / 2 for i in xs], [m["profit"] / 10000 for m in main], width=width, color=GREEN_DARK, label="净利润")
    ax.set_xticks(xs)
    ax.set_xticklabels([m["solution"] for m in main], rotation=20)
    ax.set_ylabel("万元")
    ax.set_title("(d) 营业额与净利润对比")
    place_legend(ax)
    save_figure(fig, FIG_MULTI_DASHBOARD)

    fig, axes = plt.subplots(2, 2, figsize=(15.5, 9.5))

    axes_names = [("营业额", "revenue_norm"), ("净利润", "profit_norm"), ("公平性", "fairness_norm"), ("满意度", "satisfaction_norm")]
    ax = axes[0, 0]
    for m in metrics:
        ax.scatter(m["revenue_norm"], m["satisfaction_norm"], color=PURPLE_DARK if m["is_pareto"] else BLUE_MID, s=45 if m["is_pareto"] else 20)
    ax.set_title("(a) 归一化营业额-满意度")
    ax.set_xlabel("营业额")
    ax.set_ylabel("满意度")

    ax = axes[0, 1]
    xs = list(range(len(axes_names)))
    for m in [m for m in metrics if m["is_pareto"]][:15]:
        vals = [m[key] for _label, key in axes_names]
        ax.plot(xs, vals, alpha=0.75, linewidth=1.6)
    ax.set_xticks(xs)
    ax.set_xticklabels([label for label, _key in axes_names])
    ax.set_ylim(0, 1)
    ax.set_title("(b) 帕累托方案平行坐标")

    ax = axes[1, 0]
    matrix = []
    for _label_a, key_a in axes_names:
        row = []
        vals_a = [m[key_a] for m in metrics]
        mean_a = sum(vals_a) / len(vals_a)
        for _label_b, key_b in axes_names:
            vals_b = [m[key_b] for m in metrics]
            mean_b = sum(vals_b) / len(vals_b)
            cov = sum((a - mean_a) * (b - mean_b) for a, b in zip(vals_a, vals_b))
            sa = math.sqrt(sum((a - mean_a) ** 2 for a in vals_a)) or 1.0
            sb = math.sqrt(sum((b - mean_b) ** 2 for b in vals_b)) or 1.0
            row.append(cov / (sa * sb))
        matrix.append(row)
    im = ax.imshow(matrix, cmap=cmap_from(PURPLE_LIGHT, BLUE_DARK, "tradeoff"), vmin=-1, vmax=1)
    ax.set_xticks(xs)
    ax.set_xticklabels([label for label, _key in axes_names])
    ax.set_yticks(xs)
    ax.set_yticklabels([label for label, _key in axes_names])
    ax.set_title("(c) 目标相关性热力图")
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    ax = axes[1, 1]
    ax.scatter([m["profit"] / 10000 for m in metrics], [m["fairness"] for m in metrics], c=[GREEN_DARK if m["is_pareto"] else PURPLE_MID for m in metrics], s=[45 if m["is_pareto"] else 20 for m in metrics])
    ax.set_title("(d) 净利润-公平性关系")
    ax.set_xlabel("净利润（万元）")
    ax.set_ylabel("公平性")
    save_figure(fig, FIG_PARETO_DASHBOARD)
    draw_solution_score_table(metrics)


def draw_all_figures(metrics: list[dict[str, Any]]) -> None:
    draw_objective_radar(metrics)
    draw_pareto_revenue_satisfaction(metrics)
    draw_pareto_pairwise_choices(metrics)
    draw_task3_dashboards(metrics)



def main() -> None:
    random.seed(RANDOM_SEED)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    rows = read_task2_rows()
    positions = read_positions()
    heat = read_node_heat()
    layouts = make_candidate_layouts(rows, positions)

    metrics = []
    all_details = []
    for name, layout in layouts.items():
        metric, details = evaluate_solution(name, layout, rows, positions, heat)
        metrics.append(metric)
        all_details.extend(details)

    normalize_metrics(metrics)
    assign_nsga_rank_and_crowding(metrics)
    mark_pareto(metrics)
    best = max(metrics, key=lambda m: m["combined_score"])
    compromise = dict(best)
    compromise["solution"] = "折中方案"
    metrics.append(compromise)
    normalize_metrics(metrics)
    assign_nsga_rank_and_crowding(metrics)
    mark_pareto(metrics)

    summary = write_csv_outputs(metrics, all_details)
    draw_all_figures(metrics)

    print(f"Candidates: {summary['candidate_count']}")
    print(f"Pareto count: {summary['pareto_count']}")
    print(f"Best compromise: {summary['best_compromise_solution']}")
    print(f"Best revenue: {summary['best_revenue_solution']}")
    print(f"Best satisfaction: {summary['best_satisfaction_solution']}")
    print(f"Wrote: {CANDIDATE_CSV}")
    print(f"Wrote: {STORE_DETAIL_CSV}")
    print(f"Wrote: {SUMMARY_CSV}")

if __name__ == "__main__":
    main()
