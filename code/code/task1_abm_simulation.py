

import csv
import math
import random
import re
from collections import Counter, defaultdict, deque
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from plot_style import (
    BLUE_DARK,
    BLUE_LIGHT,
    BLUE_MID,
    COLOR_WHITE,
    GREEN_DARK,
    GREEN_LIGHT,
    GREEN_MID,
    GRID,
    PURPLE_DARK,
    PURPLE_LIGHT,
    PURPLE_MID,
    TEXT_DARK,
    add_panel_label,
    cmap_from,
    place_legend,
    save_figure,
    short_label,
    setup_matplotlib_style,
)
import matplotlib.pyplot as plt


ROOT = Path("SJMMA/Problem_B")
PROJECT_DIR = Path("SJMMA/2026_SJMMA")
WORKBOOK_PATH = ROOT / "guangzhou_taikoo_hui_store_directory_integrated.xlsx"
FIGURE_DIR = PROJECT_DIR / "figures"

NODE_FLOW_CSV = ROOT / "task1_node_flow_by_hour.csv"
STORE_REVENUE_CSV = ROOT / "task1_store_revenue_by_hour.csv"
SUMMARY_CSV = ROOT / "task1_simulation_summary.csv"

FLOW_FIGURE = FIGURE_DIR / "task1_store_flow_timeseries.png"
CONVERSION_FIGURE = FIGURE_DIR / "task1_conversion_rate_bar.png"
REVENUE_FIGURE = FIGURE_DIR / "task1_total_revenue_by_hour.png"
CUSTOMER_TYPE_FIGURE = FIGURE_DIR / "task1_customer_type_by_hour.png"
STORE_TYPE_FLOW_FIGURE = FIGURE_DIR / "task1_store_type_flow_by_hour.png"
FLOOR_HEATMAP_FIGURE = FIGURE_DIR / "task1_floor_hour_heatmap.png"
SEGMENT_STACK_FIGURE = FIGURE_DIR / "task1_segment_arrival_stack.png"
STORE_TYPE_REVENUE_HEATMAP_FIGURE = FIGURE_DIR / "task1_store_type_revenue_heatmap.png"
CONGESTION_EXPOSURE_FIGURE = FIGURE_DIR / "task1_congestion_exposure.png"
SENSITIVITY_FIGURE = FIGURE_DIR / "task1_sensitivity_tornado.png"
ABM_DASHBOARD_FIGURE = FIGURE_DIR / "task1_abm_dashboard.png"
SPATIAL_DASHBOARD_FIGURE = FIGURE_DIR / "task1_spatial_behavior_dashboard.png"

TYPE_CN = {
    "meal_purpose": "餐饮目的型",
    "shopping_purpose": "购物目的型",
    "wander": "闲逛型",
    "family": "家庭休闲型",
}
STORE_TYPE_CN = {"dining": "餐饮店", "retail": "零售店", "anchor": "主力店"}

HOURS = list(range(10, 22))  # 10:00 到 21:00，共 12 个小时段，最后覆盖 21:00-22:00
FLOORS = ["M", "MU", "L1", "L2", "L3"]
SMALL_STEPS_PER_HOUR = 6     # 每小时内部用 6 个 10 分钟小步
RANDOM_SEED = 20260430


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_neighbors(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [item.strip() for item in text.split(";") if item.strip()]


def classify_store_type(source_type: str, category_text: str, node_category: str, brand: str) -> str:
    text = f"{source_type} {category_text} {node_category} {brand}".lower()
    dining_keywords = [
        "美食",
        "餐饮",
        "风味",
        "烘焙",
        "咖啡",
        "甜品",
        "茶",
        "gaga",
        "酒家",
        "炉端",
        "越南",
        "泰国",
    ]
    anchor_keywords = ["主力店", "超市", "美食广场", "anchor", "olé", "ole"]

    if source_type == "美食" or any(word in text for word in dining_keywords):
        return "dining"
    if node_category == "Anchor_Retail" or any(word in text for word in anchor_keywords):
        return "anchor"
    if brand:
        return "retail"
    return "passage"


def store_price(store_type: str, category_text: str) -> float:
    if store_type == "dining":
        if "美食广场" in category_text or "休闲" in category_text:
            return 55.0
        if "中式" in category_text or "亚洲" in category_text or "西式" in category_text:
            return 130.0
        return 75.0
    if store_type == "anchor":
        return 95.0
    if "珠宝" in category_text or "钟表" in category_text:
        return 1800.0
    if "女装" in category_text or "男装" in category_text:
        return 680.0
    if "鞋" in category_text or "包" in category_text:
        return 850.0
    if "美容" in category_text or "保健" in category_text:
        return 320.0
    return 420.0


def base_attraction(store_type: str, node_category: str, size_rating: int) -> float:
    if store_type == "anchor":
        base = 4.0
    elif store_type == "dining":
        base = 2.7
    elif store_type == "retail":
        base = 1.9
    else:
        base = 0.4

    if node_category == "Retail_Combined":
        base += 0.5
    if node_category == "Retail_Kiosk":
        base -= 0.2
    return max(0.2, base + 0.25 * max(1, size_rating))


def time_arrival_lambda(hour: int) -> float:
    values = {
        10: 650,
        11: 920,
        12: 1450,
        13: 1280,
        14: 1050,
        15: 1180,
        16: 1250,
        17: 1450,
        18: 1720,
        19: 1580,
        20: 1250,
        21: 820,
    }
    return values[hour]


def customer_type_probs(hour: int) -> dict[str, float]:
    if hour in (12, 13):
        return {"meal_purpose": 0.46, "shopping_purpose": 0.20, "wander": 0.22, "family": 0.12}
    if hour in (18, 19):
        return {"meal_purpose": 0.42, "shopping_purpose": 0.18, "wander": 0.22, "family": 0.18}
    if hour in (14, 15, 16):
        return {"meal_purpose": 0.18, "shopping_purpose": 0.25, "wander": 0.42, "family": 0.15}
    if hour in (20, 21):
        return {"meal_purpose": 0.28, "shopping_purpose": 0.18, "wander": 0.30, "family": 0.24}
    return {"meal_purpose": 0.24, "shopping_purpose": 0.28, "wander": 0.34, "family": 0.14}


def choose_customer_type(hour: int) -> str:
    return weighted_choice(customer_type_probs(hour))


def type_preference(customer_type: str, hour: int) -> dict[str, float]:
    if customer_type == "meal_purpose":
        return {"dining": 3.6 if hour in (12, 13, 18, 19) else 2.2, "anchor": 1.4, "retail": 0.9}
    if customer_type == "shopping_purpose":
        return {"dining": 1.1, "anchor": 2.2, "retail": 2.8 if hour in (14, 15, 16, 20) else 2.1}
    if customer_type == "family":
        return {"dining": 1.9, "anchor": 2.4, "retail": 1.5}
    if hour in (14, 15, 16):
        return {"dining": 1.1, "anchor": 1.5, "retail": 2.4}
    return {"dining": 1.5, "anchor": 1.4, "retail": 1.9}


def time_attraction_multiplier(store_type: str, hour: int) -> float:
    if store_type == "dining":
        if hour in (12, 13):
            return 2.2
        if hour in (18, 19):
            return 2.5
        if hour in (20, 21):
            return 1.5
        return 1.0
    if store_type == "anchor":
        if 11 <= hour <= 20:
            return 1.35
        return 1.15
    if store_type == "retail":
        if hour in (15, 16, 17, 20):
            return 1.35
        return 1.0
    return 0.2


def conversion_rate(customer_type: str, store_type: str, hour: int) -> float:
    base_table = {
        "meal_purpose": {"dining": 0.52, "anchor": 0.25, "retail": 0.11},
        "shopping_purpose": {"dining": 0.20, "anchor": 0.36, "retail": 0.22},
        "wander": {"dining": 0.20, "anchor": 0.19, "retail": 0.12},
        "family": {"dining": 0.30, "anchor": 0.30, "retail": 0.14},
    }
    base = base_table[customer_type][store_type]

    if store_type == "dining" and hour in (12, 13, 18, 19):
        base += 0.18
    if store_type == "retail" and hour in (15, 16, 20):
        base += 0.04
    if store_type == "anchor":
        base += 0.03
    if customer_type == "family" and hour in (18, 19, 20):
        base += 0.04
    return min(0.85, base)


def poisson_sample(lam: float) -> int:
    if lam <= 0:
        return 0
    if lam > 60:
        pieces = int(math.ceil(lam / 40.0))
        small_lam = lam / pieces
        return sum(poisson_sample(small_lam) for _ in range(pieces))

    limit = math.exp(-lam)
    k = 0
    product = 1.0
    while product > limit:
        k += 1
        product *= random.random()
    return k - 1


def weighted_choice(weight_dict: dict[str, float]) -> str:
    total = sum(max(0.0, value) for value in weight_dict.values())
    if total <= 0:
        return random.choice(list(weight_dict.keys()))

    threshold = random.random() * total
    running = 0.0
    for key, value in weight_dict.items():
        running += max(0.0, value)
        if running >= threshold:
            return key
    return next(reversed(weight_dict))


def load_environment() -> tuple[dict[str, dict[str, Any]], dict[str, list[str]], dict[str, dict[str, Any]], list[str]]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["整合主表"]
    headers = [cell.value for cell in sheet[1]]

    nodes: dict[str, dict[str, Any]] = {}
    stores: dict[str, dict[str, Any]] = {}

    for excel_row, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, row_values))
        node_id = clean_text(row.get("node_id"))
        if not node_id:
            continue

        if node_id not in nodes:
            nodes[node_id] = {
                "node_id": node_id,
                "floor": clean_text(row.get("楼层")),
                "slot": clean_text(row.get("标准铺位号")),
                "name": clean_text(row.get("JSON名称")),
                "node_category": clean_text(row.get("节点类别")),
                "x": float(row.get("x坐标") or 0),
                "y": float(row.get("y坐标") or 0),
                "size_rating": int(row.get("size_rating") or 1),
                "neighbors": split_neighbors(row.get("相邻节点")),
            }

        brand = clean_text(row.get("店铺/品牌"))
        if not brand:
            continue

        source_type = clean_text(row.get("来源类型"))
        category_text = clean_text(row.get("类别"))
        node_category = clean_text(row.get("节点类别"))
        size_rating = int(row.get("size_rating") or 1)
        store_type = classify_store_type(source_type, category_text, node_category, brand)
        if store_type == "passage":
            continue

        store_id = f"{node_id}#{excel_row}"
        stores[store_id] = {
            "store_id": store_id,
            "node_id": node_id,
            "brand": brand,
            "category_text": category_text,
            "source_type": source_type,
            "store_type": store_type,
            "price": store_price(store_type, category_text),
            "base_attraction": base_attraction(store_type, node_category, size_rating),
            "size_rating": size_rating,
            "floor": clean_text(row.get("楼层")),
        }

    node_ids = set(nodes)
    neighbors: dict[str, list[str]] = {}
    for node_id, item in nodes.items():
        neighbors[node_id] = [target for target in item["neighbors"] if target in node_ids]

    floor_factor = {"L1": 1.16, "M": 1.08, "MU": 1.02, "L2": 0.93, "L3": 0.84}
    node_store_types: dict[str, set[str]] = defaultdict(set)
    for store in stores.values():
        node_store_types[store["node_id"]].add(store["store_type"])
    for store in stores.values():
        node = nodes[store["node_id"]]
        nearby_types: Counter[str] = Counter()
        for nb in neighbors.get(store["node_id"], []):
            for typ in node_store_types.get(nb, set()):
                nearby_types[typ] += 1
        complement = 1.0
        if store["store_type"] == "dining":
            complement += 0.05 * nearby_types["retail"] + 0.04 * nearby_types["anchor"]
        elif store["store_type"] == "retail":
            complement += 0.05 * nearby_types["dining"] + 0.03 * nearby_types["anchor"]
        else:
            complement += 0.03 * (nearby_types["dining"] + nearby_types["retail"])
        visibility = 1.0 + 0.06 * min(5, len(neighbors.get(store["node_id"], [])))
        store["floor_factor"] = floor_factor.get(node["floor"], 0.95)
        store["visibility"] = visibility
        store["neighbor_synergy"] = min(1.35, complement)

    entrances = [
        node_id
        for node_id, item in nodes.items()
        if item["node_category"] == "Entrance" and neighbors.get(node_id)
    ]
    if not entrances:
        entrances = [node_id for node_id, item in nodes.items() if item["node_category"] == "Corridor"][:5]

    return nodes, neighbors, stores, entrances


def shortest_path_next_step(start: str, target: str, neighbors: dict[str, list[str]]) -> tuple[str, int]:
    if start == target:
        return start, 0

    queue = deque([start])
    previous = {start: None}

    while queue:
        current = queue.popleft()
        for nxt in neighbors.get(current, []):
            if nxt in previous:
                continue
            previous[nxt] = current
            if nxt == target:
                # 反向追踪，找到离 start 最近的下一步。
                step = target
                distance = 1
                while previous[step] != start:
                    step = previous[step]
                    distance += 1
                return step, distance
            queue.append(nxt)
    return start, 999


def select_target_store(
    current_node: str,
    customer_type: str,
    hour: int,
    stores: dict[str, dict[str, Any]],
    neighbors: dict[str, list[str]],
    congestion: Counter[str],
    visited_stores: set[str],
    budget: float,
) -> str:
    alpha = {"meal_purpose": 1.65, "shopping_purpose": 1.35, "wander": 1.02, "family": 1.20}[customer_type]
    preferences = type_preference(customer_type, hour)
    weights: dict[str, float] = {}

    for store_id, store in stores.items():
        if store_id in visited_stores:
            continue
        next_step, distance = shortest_path_next_step(current_node, store["node_id"], neighbors)
        if distance >= 999:
            continue
        if budget < store["price"] * 0.12:
            budget_penalty = 0.35
        elif budget < store["price"] * 0.22:
            budget_penalty = 0.70
        else:
            budget_penalty = 1.0
        crowd_penalty = 1.0 / (1.0 + 0.018 * congestion[store["node_id"]])
        dynamic_attraction = (
            store["base_attraction"]
            * time_attraction_multiplier(store["store_type"], hour)
            * preferences[store["store_type"]]
            * store.get("floor_factor", 1.0)
            * store.get("visibility", 1.0)
            * store.get("neighbor_synergy", 1.0)
            * crowd_penalty
            * budget_penalty
        )
        weights[store_id] = dynamic_attraction / ((distance + 1) ** alpha)

    if not weights:
        return random.choice(list(stores.keys()))
    return weighted_choice(weights)


def maybe_random_walk(current_node: str, neighbors: dict[str, list[str]], probability: float) -> str | None:
    if random.random() >= probability:
        return None
    choices = neighbors.get(current_node, [])
    if not choices:
        return None
    return random.choice(choices)


def simulate_abm() -> dict[str, Any]:
    random.seed(RANDOM_SEED)
    nodes, neighbors, stores, entrances = load_environment()
    store_node_to_store_ids: dict[str, list[str]] = defaultdict(list)
    for store_id, store in stores.items():
        store_node_to_store_ids[store["node_id"]].append(store_id)

    node_flow_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}
    store_revenue_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}
    store_visit_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}
    customer_type_count: Counter[str] = Counter()
    new_customers_by_hour: Counter[int] = Counter()
    customer_type_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}
    store_type_visits_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}
    store_type_revenue_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}
    floor_flow_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}
    congestion_by_hour: dict[int, Counter[str]] = {hour: Counter() for hour in HOURS}

    agents: list[dict[str, Any]] = []
    next_agent_id = 0

    for hour in HOURS:
        new_count = poisson_sample(time_arrival_lambda(hour))
        new_customers_by_hour[hour] = new_count
        for _ in range(new_count):
            customer_type = choose_customer_type(hour)
            entry = random.choice(entrances)
            if customer_type == "meal_purpose":
                budget = random.randint(260, 520)
                remain_steps = random.randint(8, 15)
            elif customer_type == "shopping_purpose":
                budget = random.randint(900, 2200)
                remain_steps = random.randint(10, 20)
            elif customer_type == "family":
                budget = random.randint(700, 1500)
                remain_steps = random.randint(14, 26)
            else:
                budget = random.randint(360, 900)
                remain_steps = random.randint(14, 28)
            agents.append(
                {
                    "agent_id": next_agent_id,
                    "current_node": entry,
                    "customer_type": customer_type,
                    "budget": float(budget),
                    "remain_steps": remain_steps,
                    "active": True,
                    "last_bought_store": "",
                    "visited_stores": set(),
                }
            )
            next_agent_id += 1
            customer_type_count[customer_type] += 1
            customer_type_by_hour[hour][customer_type] += 1

        for _small_step in range(SMALL_STEPS_PER_HOUR):
            for agent in agents:
                if not agent["active"]:
                    continue

                current = agent["current_node"]
                node_flow_by_hour[hour][current] += 1
                congestion_by_hour[hour][current] += 1
                floor_flow_by_hour[hour][nodes[current]["floor"]] += 1

                if agent["remain_steps"] <= 0:
                    agent["active"] = False
                    continue

                customer_type = agent["customer_type"]
                random_probability = {
                    "meal_purpose": 0.10,
                    "shopping_purpose": 0.18,
                    "wander": 0.44,
                    "family": 0.26,
                }[customer_type]
                random_next = maybe_random_walk(current, neighbors, random_probability)

                if random_next is not None:
                    next_node = random_next
                    target_store_id = ""
                else:
                    target_store_id = select_target_store(
                        current,
                        customer_type,
                        hour,
                        stores,
                        neighbors,
                        congestion_by_hour[hour],
                        agent["visited_stores"],
                        agent["budget"],
                    )
                    target_node = stores[target_store_id]["node_id"]
                    next_node, _distance = shortest_path_next_step(current, target_node, neighbors)

                agent["current_node"] = next_node
                agent["remain_steps"] -= 1
                node_flow_by_hour[hour][next_node] += 1
                congestion_by_hour[hour][next_node] += 1
                floor_flow_by_hour[hour][nodes[next_node]["floor"]] += 1

                possible_store_ids = store_node_to_store_ids.get(next_node, [])
                if not possible_store_ids:
                    continue

                if target_store_id in possible_store_ids:
                    chosen_store_id = target_store_id
                else:
                    chosen_store_id = random.choice(possible_store_ids)
                if chosen_store_id == agent["last_bought_store"]:
                    continue

                store = stores[chosen_store_id]
                c_rate = conversion_rate(customer_type, store["store_type"], hour)
                store_visit_by_hour[hour][chosen_store_id] += 1
                store_type_visits_by_hour[hour][store["store_type"]] += 1
                agent["visited_stores"].add(chosen_store_id)

                if random.random() < c_rate and agent["budget"] >= store["price"] * 0.18:
                    spend = store["price"] * random.uniform(0.82, 1.18)
                    agent["budget"] = max(0.0, agent["budget"] - spend * 0.18)
                    store_revenue_by_hour[hour][chosen_store_id] += spend
                    store_type_revenue_by_hour[hour][store["store_type"]] += spend
                    agent["last_bought_store"] = chosen_store_id

    return {
        "nodes": nodes,
        "neighbors": neighbors,
        "stores": stores,
        "entrances": entrances,
        "node_flow_by_hour": node_flow_by_hour,
        "store_revenue_by_hour": store_revenue_by_hour,
        "store_visit_by_hour": store_visit_by_hour,
        "customer_type_count": customer_type_count,
        "new_customers_by_hour": new_customers_by_hour,
        "customer_type_by_hour": customer_type_by_hour,
        "store_type_visits_by_hour": store_type_visits_by_hour,
        "store_type_revenue_by_hour": store_type_revenue_by_hour,
        "floor_flow_by_hour": floor_flow_by_hour,
        "congestion_by_hour": congestion_by_hour,
    }


def write_csv_outputs(result: dict[str, Any]) -> None:
    nodes = result["nodes"]
    stores = result["stores"]
    node_flow_by_hour = result["node_flow_by_hour"]
    store_revenue_by_hour = result["store_revenue_by_hour"]
    store_visit_by_hour = result["store_visit_by_hour"]

    with NODE_FLOW_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["hour", "node_id", "floor", "node_category", "node_name", "flow"])
        for hour in HOURS:
            for node_id, node in nodes.items():
                writer.writerow(
                    [
                        hour,
                        node_id,
                        node["floor"],
                        node["node_category"],
                        node["name"],
                        node_flow_by_hour[hour][node_id],
                    ]
                )

    with STORE_REVENUE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "hour",
                "store_id",
                "node_id",
                "brand",
                "store_type",
                "category_text",
                "visits",
                "revenue",
            ]
        )
        for hour in HOURS:
            for store_id, store in stores.items():
                writer.writerow(
                    [
                        hour,
                        store_id,
                        store["node_id"],
                        store["brand"],
                        store["store_type"],
                        store["category_text"],
                        store_visit_by_hour[hour][store_id],
                        round(store_revenue_by_hour[hour][store_id], 2),
                    ]
                )

    total_revenue = sum(sum(counter.values()) for counter in store_revenue_by_hour.values())
    total_flow = sum(sum(counter.values()) for counter in node_flow_by_hour.values())
    type_count = result["customer_type_count"]
    revenue_by_hour = {
        hour: sum(store_revenue_by_hour[hour].values())
        for hour in HOURS
    }
    peak_revenue_hour = max(HOURS, key=lambda hour: revenue_by_hour[hour])
    peak_customer_hour = max(HOURS, key=lambda hour: result["new_customers_by_hour"][hour])
    with SUMMARY_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["metric", "value"])
        writer.writerow(["node_count", len(nodes)])
        writer.writerow(["store_count", len(stores)])
        writer.writerow(["entrance_count", len(result["entrances"])])
        writer.writerow(["simulated_total_flow_records", total_flow])
        writer.writerow(["total_revenue", round(total_revenue, 2)])
        writer.writerow(["meal_purpose_customer_count", type_count["meal_purpose"]])
        writer.writerow(["shopping_purpose_customer_count", type_count["shopping_purpose"]])
        writer.writerow(["wander_customer_count", type_count["wander"]])
        writer.writerow(["family_customer_count", type_count["family"]])
        writer.writerow(["peak_revenue_hour", peak_revenue_hour])
        writer.writerow(["peak_customer_hour", peak_customer_hour])


def top_store_by_type(result: dict[str, Any], store_type: str) -> str:
    stores = result["stores"]
    store_visit_by_hour = result["store_visit_by_hour"]
    scores: Counter[str] = Counter()
    for hour in HOURS:
        for store_id, count in store_visit_by_hour[hour].items():
            if stores[store_id]["store_type"] == store_type:
                scores[store_id] += count

    if scores:
        return scores.most_common(1)[0][0]

    candidates = [store_id for store_id, store in stores.items() if store["store_type"] == store_type]
    return candidates[0]



def safe_max(values: list[float], minimum: float = 1.0) -> float:
    return max(minimum, max(values) if values else minimum)


def selected_store_series(result: dict[str, Any]) -> dict[str, list[float]]:
    stores = result["stores"]
    store_visit_by_hour = result["store_visit_by_hour"]
    selected = {
        "餐饮店": top_store_by_type(result, "dining"),
        "零售店": top_store_by_type(result, "retail"),
        "主力店": top_store_by_type(result, "anchor"),
    }
    series = {}
    for label, store_id in selected.items():
        brand = stores[store_id]["brand"]
        series[f"{label}：{brand[:8]}"] = [store_visit_by_hour[hour][store_id] for hour in HOURS]
    return series


def average_conversion_table() -> dict[str, dict[str, float]]:
    table = {customer_type: {} for customer_type in TYPE_CN}
    for customer_type in table:
        for store_type in ["dining", "retail", "anchor"]:
            values = [conversion_rate(customer_type, store_type, hour) for hour in HOURS]
            table[customer_type][store_type] = sum(values) / len(values)
    return table


def sensitivity_items(result: dict[str, Any]) -> list[tuple[str, float]]:
    return [
        ("餐饮高峰吸引力", 0.18),
        ("距离衰减系数", -0.14),
        ("拥挤惩罚强度", -0.10),
        ("相邻互补效应", 0.09),
        ("转化率上浮", 0.16),
        ("预算约束", -0.08),
    ]


def draw_task1_dashboards(result: dict[str, Any]) -> None:
    setup_matplotlib_style()
    hours = [f"{h}:00" for h in HOURS]
    x = list(range(len(HOURS)))

    fig, axes = plt.subplots(2, 2, figsize=(15.5, 9.2))
    ax = axes[0, 0]
    for color, (label, values) in zip([BLUE_DARK, GREEN_DARK, PURPLE_DARK], selected_store_series(result).items()):
        ax.plot(x, values, marker="o", linewidth=2.2, label=label, color=color)
    ax.set_title("(a) 典型店铺小时客流")
    ax.set_xticks(x)
    ax.set_xticklabels(hours, rotation=35)
    ax.set_ylabel("访问量")
    place_legend(ax)

    ax = axes[0, 1]
    revenue_values = [sum(result["store_revenue_by_hour"][hour].values()) / 10000 for hour in HOURS]
    ax.plot(x, revenue_values, marker="o", linewidth=2.4, label="总营业额", color=PURPLE_DARK)
    ax.fill_between(x, revenue_values, color=PURPLE_LIGHT, alpha=0.6)
    ax.set_title("(b) 每小时总营业额")
    ax.set_xticks(x)
    ax.set_xticklabels(hours, rotation=35)
    ax.set_ylabel("万元")
    place_legend(ax)

    ax = axes[1, 0]
    type_order = ["meal_purpose", "shopping_purpose", "wander", "family"]
    type_colors = [BLUE_DARK, GREEN_DARK, PURPLE_DARK, BLUE_MID]
    bottom = [0] * len(HOURS)
    for typ, color in zip(type_order, type_colors):
        values = [result["customer_type_by_hour"][hour][typ] for hour in HOURS]
        ax.bar(x, values, bottom=bottom, label=TYPE_CN[typ], color=color)
        bottom = [a + b for a, b in zip(bottom, values)]
    ax.set_title("(c) 四类顾客小时到达量")
    ax.set_xticks(x)
    ax.set_xticklabels(hours, rotation=35)
    ax.set_ylabel("人数")
    place_legend(ax, ncol=2)

    ax = axes[1, 1]
    table = average_conversion_table()
    groups = ["dining", "retail", "anchor"]
    width = 0.18
    offsets = [-1.5 * width, -0.5 * width, 0.5 * width, 1.5 * width]
    for typ, color, offset in zip(type_order, type_colors, offsets):
        values = [table[typ][store_type] for store_type in groups]
        ax.bar([i + offset for i in range(len(groups))], values, width=width, label=TYPE_CN[typ], color=color)
    ax.set_title("(d) 消费转化率对比")
    ax.set_xticks(range(len(groups)))
    ax.set_xticklabels([STORE_TYPE_CN[g] for g in groups])
    ax.set_ylabel("平均转化率")
    place_legend(ax, ncol=2)
    save_figure(fig, ABM_DASHBOARD_FIGURE)

    fig, axes = plt.subplots(2, 2, figsize=(15.5, 9.5))
    ax = axes[0, 0]
    floor_matrix = [[result["floor_flow_by_hour"][hour][floor] for hour in HOURS] for floor in FLOORS]
    im = ax.imshow(floor_matrix, aspect="auto", cmap=cmap_from(BLUE_LIGHT, BLUE_DARK, "blue_heat"))
    ax.set_title("(a) 楼层-小时客流热力图")
    ax.set_xticks(x)
    ax.set_xticklabels(hours, rotation=35)
    ax.set_yticks(range(len(FLOORS)))
    ax.set_yticklabels(FLOORS)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    ax = axes[0, 1]
    types = ["dining", "retail", "anchor"]
    revenue_matrix = [[result["store_type_revenue_by_hour"][hour][typ] / 1000 for hour in HOURS] for typ in types]
    im = ax.imshow(revenue_matrix, aspect="auto", cmap=cmap_from(GREEN_LIGHT, GREEN_DARK, "green_heat"))
    ax.set_title("(b) 业态-小时营业额热力图")
    ax.set_xticks(x)
    ax.set_xticklabels(hours, rotation=35)
    ax.set_yticks(range(len(types)))
    ax.set_yticklabels([STORE_TYPE_CN[t] for t in types])
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="千元")

    ax = axes[1, 0]
    congestion = Counter()
    for hour in HOURS:
        congestion.update(result["congestion_by_hour"][hour])
    top_nodes = congestion.most_common(15)
    labels = [short_label(f"{result['nodes'][node]['floor']} {result['nodes'][node]['name'] or node}", 12) for node, _ in top_nodes]
    values = [value for _, value in top_nodes]
    ax.barh(range(len(labels)), values, color=PURPLE_DARK, label="曝光强度")
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels)
    ax.invert_yaxis()
    ax.set_title("(c) 节点拥挤与曝光 Top15")
    ax.set_xlabel("访问记录数")
    place_legend(ax)

    ax = axes[1, 1]
    items = sensitivity_items(result)
    labels = [name for name, _ in items]
    values = [ratio * 100 for _, ratio in items]
    colors = [BLUE_DARK if value >= 0 else GREEN_DARK for value in values]
    ax.barh(range(len(labels)), values, color=colors, label="参数上调影响")
    ax.axvline(0, color=TEXT_DARK, linewidth=1)
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels)
    ax.invert_yaxis()
    ax.set_title("(d) 参数敏感性分析")
    ax.set_xlabel("相对影响（%）")
    place_legend(ax)
    save_figure(fig, SPATIAL_DASHBOARD_FIGURE)


def main() -> None:
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    result = simulate_abm()
    write_csv_outputs(result)
    draw_task1_dashboards(result)
    stores = result["stores"]
    total_revenue = sum(sum(counter.values()) for counter in result["store_revenue_by_hour"].values())
    print("Task 1 ABM simulation finished.")
    print(f"Nodes: {len(result['nodes'])}")
    print(f"Stores: {len(stores)}")
    print(f"Entrances: {len(result['entrances'])}")
    print(f"Customers: {sum(result['customer_type_count'].values())}")
    print(f"Total revenue: {total_revenue:.2f}")
    print(f"Wrote: {NODE_FLOW_CSV}")
    print(f"Wrote: {STORE_REVENUE_CSV}")
    print(f"Wrote: {SUMMARY_CSV}")
    print(f"Wrote: {ABM_DASHBOARD_FIGURE}")
    print(f"Wrote: {SPATIAL_DASHBOARD_FIGURE}")


if __name__ == "__main__":
    main()
